# Forked from Gist: https://gist.github.com/foulegg/ed172f4c6bc42852567049c969b41049

import re
import sys
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY, FORMAT_NUMBER_00

class Transaction:
    def __init__(self, fund_name, fund_type, isin, folio, txn_data):
        self.fund_name = fund_name
        self.fund_type = fund_type

        # We only need ISIN for equity funds
        self.isin = isin if fund_type == 'MF (Equity)' else None

        self.folio = folio

        serial_no, units, purchase_date, purchase_value, \
        acquisition_value, jan31_value, redemption_date, \
        redemption_value, stcg, ltcg \
            = txn_data

        self.serial_no = int(serial_no)
        self.units = float(units)
        self.purchase_date = datetime.strptime(purchase_date, '%b %d, %Y').strftime('%d/%m/%Y')
        self.purchase_value = float(purchase_value)
        self.acquisition_value = float(acquisition_value)

        try:
            self.jan31_value = float(jan31_value)
        except:
            self.jan31_value = None

        self.redemption_date = datetime.strptime(redemption_date, '%b %d, %Y').strftime('%d/%m/%Y')
        self.redemption_value = float(redemption_value)

        # Take out "," from CG numbers
        try:
            self.stcg = float(stcg.replace(',', ''))
        except:
            self.stcg = None

        try:
            self.ltcg = float(ltcg.replace(',', ''))
        except:
            self.ltcg = None

    def get_data(self):
        return (
            self.fund_type,
            self.isin,
            self.fund_name,
            self.units,
            self.purchase_date,
            self.purchase_value,
            self.redemption_date,
            self.redemption_value / self.units if self.redemption_value else None,
            self.jan31_value / self.units if self.jan31_value else None,
            0.0 # Set "Transfer expenses (stamp duty, brokerage, etc.) column to 0
        )

# Reads all transactions from the capital gains report from Kuvera and returns the parsed data
# in a tuple of (transactions, total_stcg, total_ltcg)
def read_transactions(capital_gains_xls_file):
    fund_name_pattern = re.compile('(.*)\[ISIN')
    isin_pattern = re.compile('\[ISIN: (.*)\]')
    folio_pattern = re.compile('Folio No: (.*)')

    with open(capital_gains_xls_file) as f:
        data = f.read()
        soup = BeautifulSoup(data, 'html.parser')

    # The second table contains capital gains data
    capital_gains_table = soup.select('table')[1]
    capital_gains_tbody = capital_gains_table.select('tbody')[0]

    current_fund_name = None
    current_fund_type = None
    current_isin = None
    current_folio = None

    equity_subtotal = None
    debt_subtotal = None
    total_stcg = None
    total_ltcg = None
    all_transactions = []

    fund_types = {
        'Equity': 'MF (Equity)',
        'Others': 'MF (Equity)', # Index funds are marked "Others" in Kuvera's report, for some reason
        'Debt': 'MF (Other than Equity)'
    }

    for row in capital_gains_tbody.select('tr'):
        columns = row.select('td')
        if len(columns) == 1:
            # This row contains either the Fund name with ISIN or the Folio number
            column_data = columns[0].string
            isin_match = isin_pattern.search(column_data)
            folio_match = folio_pattern.search(column_data)
            if isin_match:
                current_isin = isin_match.group(1)
            elif folio_match:
                current_folio = folio_match.group(1)

            # Extract fund name
            fund_name_match = fund_name_pattern.search(column_data)
            if fund_name_match:
                current_fund_name = fund_name_match.group(1)

            for fund_type in fund_types:
                if fund_type in column_data:
                    current_fund_type = fund_types[fund_type]
                    break
        elif len(columns) == 10:
            # This row contains a transaction
            txn_data = [col.string for col in columns]
            transaction = Transaction(current_fund_name, current_fund_type, current_isin, current_folio, txn_data)
            all_transactions.append(transaction)
        elif len(columns) == 8:
            # Summary of transactions for a fund, we can skip these rows
            pass
        elif len(columns) == 3:
            # Rows with subtotal and total
            row_title = columns[0].string
            if row_title == 'Total':
                total_stcg = float(columns[1].string.split()[1].replace(',', ''))
                total_ltcg = float(columns[2].string.split()[1].replace(',', ''))

    return (all_transactions, total_stcg, total_ltcg)

def write_capital_gains_report(all_transactions, output_xlsx_file, cleartax_template_xlsx_file):

    wb = load_workbook(cleartax_template_xlsx_file)

    # We only populate data in the Mutual Fund Sheet, which is the third one in Cleartax template (as of 14 Oct 2020)
    mf_sheet = wb.worksheets[2]

    # Columns 9, 11 and 13 are computed from the other columns
    columns_to_update = [1, 2, 3, 4, 5, 6, 7, 8, 10, 12]
    column_types = [
        'string',
        'string',
        'string',
        'float',
        'date',
        'float',
        'date',
        'float',
        'float',
        'string'
    ]
    for i in range(len(all_transactions)):
        txn_data = all_transactions[i].get_data()
        for (val, j, col_type) in zip(txn_data, columns_to_update, column_types):
            cell = mf_sheet.cell(row=i+2, column=j)
            if val is not None:
                cell.value = val
                if col_type == 'float':
                    cell.number_format = FORMAT_NUMBER_00
                elif col_type == 'date':
                    cell.number_format = FORMAT_DATE_DDMMYY

    wb.save(output_xlsx_file)

def prepare(capital_gains_xls_file, cleartax_template_xlsx_file, output_xlsx_file):

    all_transactions, total_stcg, total_ltcg = read_transactions(capital_gains_xls_file)

    stcg_sum = sum([txn.stcg for txn in all_transactions if txn.stcg is not None])
    ltcg_sum = sum([txn.ltcg for txn in all_transactions if txn.ltcg is not None])
    print('Calculated Sum of all STCG across all transactions:', stcg_sum)
    print('Total STCG from Kuvera report:', total_stcg)
    print('Calculated Sum of all LTCG across all transactions:', ltcg_sum)
    print('Total LTCG from Kuvera report:', total_ltcg)

    write_capital_gains_report(all_transactions, output_xlsx_file, cleartax_template_xlsx_file)

if __name__ == '__main__':
    prepare(sys.argv[1], sys.argv[2], sys.argv[3])

